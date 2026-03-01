"""Report generation — Daily PDF, Weekly PDF, Security PDF, Full Excel.

This module generates downloadable reports from the analytics data stored in
SQLite.  Each generator function queries the DB, builds a formatted document,
writes it to a temporary file, and returns the file path for the FastAPI
endpoint to serve.

Report types:
  - daily_pdf     → 24-hour KPIs, outcome distribution, top intents, security summary
  - weekly_pdf    → 7-day KPIs with daily breakdown table
  - security_pdf  → Full list of security events (injection blocks, PII detections)
  - full_excel    → 4-sheet workbook: Summary, Chat Details, Security Events, Intent Breakdown
"""

import tempfile
from datetime import datetime, timezone

from mcp_server.tools.analytics_logger import (
    get_analytics_summary,
    get_intent_distribution,
    get_security_events,
    get_sessions_list,
    get_time_series,
)


def _make_temp_path(suffix: str) -> str:
    """Create a temporary file path for report output.

    The file is NOT created here — just a unique path is generated.
    The caller (reportlab / openpyxl) will write to this path.
    """
    return tempfile.mktemp(suffix=suffix, prefix="insurechat_report_")


def _build_styled_table(data, col_widths, header_color,
                        font_size=10, row_alt_color=None, extra_styles=None):
    """Build a ReportLab Table with standard InsureChat report styling."""
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_color)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]
    if row_alt_color:
        style_cmds.append(
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor(row_alt_color)])
        )
    if extra_styles:
        style_cmds.extend(extra_styles)

    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle(style_cmds))
    return table


def _init_report(title):
    """Create a PDF document with standard InsureChat report header.

    Returns (file_path, doc, styles, elements).
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    path = _make_temp_path(".pdf")
    doc = SimpleDocTemplate(path, pagesize=A4)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=18)
    elements = [
        Paragraph(title, title_style),
        Spacer(1, 12),
        Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            styles["Normal"],
        ),
        Spacer(1, 24),
    ]
    return path, doc, styles, elements


def _add_daily_kpi_table(elements, styles, summary):
    """Add the daily KPI summary table to report elements."""
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Spacer

    elements.append(Paragraph("Key Performance Indicators (Last 24 hours)", styles["Heading2"]))
    kpi_data = [
        ["Metric", "Value"],
        ["Total Sessions", str(summary.get("total_sessions", 0))],
        ["Total Messages", str(summary.get("total_messages", 0))],
        ["Success Rate", f"{summary.get('success_rate', 0):.1f}%"],
        ["Avg Confidence", f"{summary.get('avg_confidence', 0):.3f}"],
        ["Avg Response Time", f"{summary.get('avg_response_time_ms', 0):.0f}ms"],
        ["Security Events", str(summary.get("security_events", 0))],
        ["CSAT Score", f"{summary.get('csat', 3.0):.1f}/5.0"],
    ]
    elements.append(_build_styled_table(
        kpi_data, [3 * inch, 2 * inch], "#2196F3",
        extra_styles=[
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ],
        row_alt_color="#F5F5F5",
    ))
    elements.append(Spacer(1, 24))


def _add_outcome_table(elements, styles, summary):
    """Add outcome distribution table to report elements."""
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Spacer

    elements.append(Paragraph("Outcome Distribution", styles["Heading2"]))
    outcome_data = [
        ["Outcome", "Count"],
        ["Success", str(summary.get("success_count", 0))],
        ["Partial", str(summary.get("partial_count", 0))],
        ["Failure", str(summary.get("failure_count", 0))],
        ["Security Flagged", str(summary.get("flagged_count", 0))],
    ]
    elements.append(_build_styled_table(outcome_data, [3 * inch, 2 * inch], "#4CAF50"))
    elements.append(Spacer(1, 24))


def _add_intent_table(elements, styles, days):
    """Add top intents table to report elements."""
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Spacer

    elements.append(Paragraph("Top Intents", styles["Heading2"]))
    intents = get_intent_distribution(days=days)
    if intents:
        intent_data = [["Intent", "Count"]]
        for i in intents[:10]:
            intent_data.append([i["intent"], str(i["count"])])
        elements.append(_build_styled_table(intent_data, [3 * inch, 2 * inch], "#FF9800"))
    else:
        elements.append(Paragraph("No intent data for the last 24 hours.", styles["Normal"]))
    elements.append(Spacer(1, 24))


def _add_security_summary_table(elements, styles, limit):
    """Add security events summary table to report elements."""
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph

    elements.append(Paragraph("Security Summary", styles["Heading2"]))
    sec_events = get_security_events(limit=limit)
    if sec_events:
        sec_data = [["Timestamp", "Type", "Severity", "Chatbot"]]
        for ev in sec_events[:limit]:
            sec_data.append([
                ev.get("timestamp", "")[:19],
                ev.get("event_type", ""),
                ev.get("severity", ""),
                ev.get("chatbot_type", ""),
            ])
        elements.append(_build_styled_table(
            sec_data,
            [2.5 * inch, 1.5 * inch, 1 * inch, 1 * inch],
            "#F44336", font_size=8,
        ))
    else:
        elements.append(Paragraph("No security events in the last 24 hours.", styles["Normal"]))


def generate_daily_pdf() -> str:
    """Generate a daily summary PDF report.

    Includes: KPI table (last 24 h), outcome distribution, top 10 intents,
    and up to 10 recent security events.
    Returns the file path to the generated PDF.
    """
    path, doc, styles, elements = _init_report("InsureChat v3.0 \u2014 Daily Summary Report")
    summary = get_analytics_summary(days=1)
    _add_daily_kpi_table(elements, styles, summary)
    _add_outcome_table(elements, styles, summary)
    _add_intent_table(elements, styles, days=1)
    _add_security_summary_table(elements, styles, limit=10)
    doc.build(elements)
    return path


def generate_weekly_pdf() -> str:
    """Generate a weekly trends PDF report.

    Includes 7-day KPI summary and a day-by-day breakdown table showing
    message volumes, average confidence, and response times.
    Returns the file path to the generated PDF.
    """
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Spacer

    path, doc, styles, elements = _init_report("InsureChat v3.0 \u2014 Weekly Trends Report")

    # --- 7-day KPI summary table ---
    elements.append(Paragraph("7-Day KPI Summary", styles["Heading2"]))
    summary = get_analytics_summary(days=7)
    kpi_data = [
        ["Metric", "Value"],
        ["Total Sessions", str(summary.get("total_sessions", 0))],
        ["Success Rate", f"{summary.get('success_rate', 0):.1f}%"],
        ["Avg Response Time", f"{summary.get('avg_response_time_ms', 0):.0f}ms"],
        ["Security Events", str(summary.get("security_events", 0))],
    ]
    elements.append(_build_styled_table(kpi_data, [3 * inch, 2 * inch], "#2196F3"))
    elements.append(Spacer(1, 24))

    # --- Daily breakdown table: one row per day ---
    elements.append(Paragraph("Daily Breakdown", styles["Heading2"]))
    ts = get_time_series(days=7)
    if ts:
        ts_data = [["Date", "Messages", "Avg Confidence", "Avg Response Time (ms)"]]
        for row in ts:
            ts_data.append([
                row.get("date", ""),
                str(row.get("total_messages", 0)),
                f"{row.get('avg_confidence', 0):.3f}",
                f"{row.get('avg_response_time_ms', 0):.0f}",
            ])
        elements.append(_build_styled_table(
            ts_data,
            [1.5 * inch, 1.2 * inch, 1.5 * inch, 2 * inch],
            "#4CAF50", font_size=9,
        ))
    else:
        elements.append(Paragraph("No data for the last 7 days.", styles["Normal"]))

    doc.build(elements)
    return path


def generate_security_pdf() -> str:
    """Generate a security incidents PDF report.

    Lists up to 100 recent security events (injection blocks, PII detections)
    with timestamps, event type, severity, chatbot, and truncated details.
    Returns the file path to the generated PDF.
    """
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Spacer

    path, doc, styles, elements = _init_report("InsureChat v3.0 \u2014 Security Incidents Report")

    events = get_security_events(limit=100)

    if events:
        elements.append(Paragraph(f"Total Security Events: {len(events)}", styles["Heading2"]))
        elements.append(Spacer(1, 12))

        sec_data = [["Timestamp", "Event Type", "Severity", "Chatbot", "Details"]]
        for ev in events:
            sec_data.append([
                ev.get("timestamp", "")[:19],
                ev.get("event_type", ""),
                ev.get("severity", ""),
                ev.get("chatbot_type", ""),
                str(ev.get("details", ""))[:50],
            ])
        elements.append(_build_styled_table(
            sec_data,
            [1.5 * inch, 1 * inch, 0.8 * inch, 0.8 * inch, 2 * inch],
            "#F44336", font_size=7, row_alt_color="#FFF3F0",
        ))
    else:
        elements.append(Paragraph("No security events recorded.", styles["Heading2"]))

    doc.build(elements)
    return path


def _init_excel_workbook():
    """Create an openpyxl Workbook with shared header styling.

    Returns (workbook, style_header_func).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")

    def style_header(ws, cols):
        """Apply header font, fill, and alignment to the first row."""
        for col_idx, header in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    return wb, style_header


def _add_summary_sheet(wb, style_header):
    """Add Summary sheet — 30-day aggregated KPIs as key-value rows."""
    ws = wb.active
    ws.title = "Summary"
    summary = get_analytics_summary(days=30)
    style_header(ws, ["Metric", "Value"])
    for i, (k, v) in enumerate(summary.items(), 2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=str(v))
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20


def _add_chat_details_sheet(wb, style_header):
    """Add Chat Details sheet — one row per session, up to 500."""
    ws = wb.create_sheet("Chat Details")
    sessions = get_sessions_list(limit=500)
    headers = [
        "session_id", "chatbot_type", "started_at", "outcome",
        "total_messages", "avg_confidence", "avg_response_time_ms",
    ]
    style_header(ws, headers)
    for i, sess in enumerate(sessions, 2):
        for j, h in enumerate(headers, 1):
            ws.cell(row=i, column=j, value=str(sess.get(h, "")))
    for j in range(len(headers)):
        ws.column_dimensions[chr(65 + j)].width = 20


def _add_security_events_sheet(wb, style_header):
    """Add Security Events sheet — one row per event, up to 500."""
    import json

    ws = wb.create_sheet("Security Events")
    sec_events = get_security_events(limit=500)
    headers = ["timestamp", "event_type", "severity", "chatbot_type", "details"]
    style_header(ws, headers)
    for i, ev in enumerate(sec_events, 2):
        for j, h in enumerate(headers, 1):
            val = ev.get(h, "")
            if isinstance(val, dict):
                val = json.dumps(val)
            ws.cell(row=i, column=j, value=str(val)[:200])
    for j in range(len(headers)):
        ws.column_dimensions[chr(65 + j)].width = 22


def _add_intent_breakdown_sheet(wb, style_header):
    """Add Intent Breakdown sheet — intent name + count over 30 days."""
    ws = wb.create_sheet("Intent Breakdown")
    intents = get_intent_distribution(days=30)
    style_header(ws, ["Intent", "Count"])
    for i, item in enumerate(intents, 2):
        ws.cell(row=i, column=1, value=item.get("intent", ""))
        ws.cell(row=i, column=2, value=item.get("count", 0))
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15


def generate_full_excel() -> str:
    """Generate a comprehensive Excel workbook with 4 sheets.

    Sheets:
      1. Summary          — 30-day aggregated KPI key-value pairs
      2. Chat Details      — Up to 500 recent sessions with metadata
      3. Security Events   — Up to 500 recent security events
      4. Intent Breakdown  — Intent name + count over last 30 days

    Returns the file path to the generated .xlsx file.
    """
    path = _make_temp_path(".xlsx")
    wb, style_header = _init_excel_workbook()
    _add_summary_sheet(wb, style_header)
    _add_chat_details_sheet(wb, style_header)
    _add_security_events_sheet(wb, style_header)
    _add_intent_breakdown_sheet(wb, style_header)
    wb.save(path)
    return path
